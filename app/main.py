from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sqlite3
import re

from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, PlainTextResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
try:
    from docx import Document
except Exception:  # python-docx may be installed after update
    Document = None
from sqlalchemy import inspect, or_, text
from sqlalchemy.orm import Session

from .database import Base, engine, get_db
from .models import BackupObject, ChangeLog, DataProtectionPolicy, HardwareSupport, SoftwareLicense, User
from .settings import get_settings
from .auth import (
    ROLE_LABELS, SESSION_COOKIE, create_initial_admin_if_needed, create_session_token,
    get_current_user, hash_password, require_api_key, require_login, require_role,
    user_can, verify_password,
)

settings = get_settings()
app = FastAPI(title=settings.app_name)
BASE_DIR = Path(__file__).resolve().parent
Base.metadata.create_all(bind=engine)


def ensure_schema() -> None:
    """Мини-миграция для существующей SQLite-базы после обновления проекта."""
    inspector = inspect(engine)
    tables = inspector.get_table_names()
    if "software_licenses" in tables:
        columns = {column["name"] for column in inspector.get_columns("software_licenses")}
        if "location" not in columns:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE software_licenses ADD COLUMN location VARCHAR(255)"))


ensure_schema()
with next(get_db()) as _db:
    create_initial_admin_if_needed(_db)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


BACKUP_FILENAME_RE = re.compile(r"^license_monitor_backup_\d{8}_\d{6}_[A-Za-z0-9_.-]+\.db$")


def get_sqlite_db_path() -> Path:
    """Возвращает путь к SQLite-файлу базы. Backup UI рассчитан именно на SQLite."""
    db_url = settings.database_url
    if not db_url.startswith("sqlite:///"):
        raise HTTPException(status_code=400, detail="UI backup supports only SQLite database_url")
    raw_path = db_url.replace("sqlite:///", "", 1)
    db_path = Path(raw_path)
    if not db_path.is_absolute():
        db_path = Path.cwd() / db_path
    return db_path.resolve()


def get_backup_dir() -> Path:
    backup_dir = Path(settings.backup_dir)
    if not backup_dir.is_absolute():
        backup_dir = Path.cwd() / backup_dir
    backup_dir.mkdir(parents=True, exist_ok=True)
    return backup_dir.resolve()


def list_db_backups() -> list[dict]:
    backup_dir = get_backup_dir()
    rows = []
    for path in sorted(backup_dir.glob("license_monitor_backup_*.db"), reverse=True):
        stat = path.stat()
        rows.append({
            "name": path.name,
            "size_bytes": stat.st_size,
            "size_mb": round(stat.st_size / 1024 / 1024, 2),
            "created_at": datetime.fromtimestamp(stat.st_mtime),
        })
    return rows


def validate_backup_filename(filename: str) -> Path:
    if not BACKUP_FILENAME_RE.match(filename):
        raise HTTPException(status_code=400, detail="Invalid backup filename")
    path = get_backup_dir() / filename
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Backup not found")
    return path


def create_sqlite_backup(username: str) -> Path:
    db_path = get_sqlite_db_path()
    if not db_path.exists():
        raise HTTPException(status_code=404, detail=f"Database file not found: {db_path}")
    safe_user = re.sub(r"[^A-Za-z0-9_.-]", "_", username or "system")[:40]
    filename = f"license_monitor_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_user}.db"
    destination = get_backup_dir() / filename
    source_conn = sqlite3.connect(str(db_path))
    try:
        backup_conn = sqlite3.connect(str(destination))
        try:
            source_conn.backup(backup_conn)
        finally:
            backup_conn.close()
    finally:
        source_conn.close()
    return destination


def parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip()[:10])


def parse_int(value, default=None):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except Exception:
        return default


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "да", "истина", "on"}


def calc_status(days_left: int) -> str:
    if days_left < 0:
        return "expired"
    if days_left <= 7:
        return "urgent"
    if days_left <= 30:
        return "critical"
    if days_left <= 60:
        return "warning"
    return "active"


def update_status(item) -> None:
    item.status = calc_status(item.days_left(date.today()))


def log_change(db: Session, section: str, record_id: int | None, action: str, description: str, actor: str = "web") -> None:
    db.add(ChangeLog(section=section, record_id=record_id, action=action, description=description, actor=actor))


def hardware_query(db: Session, q: str | None = None, status: str | None = None, responsible: str | None = None):
    query = db.query(HardwareSupport)
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(
            HardwareSupport.dzo.ilike(pattern), HardwareSupport.serial_number.ilike(pattern),
            HardwareSupport.equipment_model.ilike(pattern), HardwareSupport.contract_number.ilike(pattern),
            HardwareSupport.contractor.ilike(pattern), HardwareSupport.responsible.ilike(pattern),
            HardwareSupport.comment.ilike(pattern),
        ))
    if status:
        query = query.filter(HardwareSupport.status == status)
    if responsible:
        query = query.filter(HardwareSupport.responsible == responsible)
    items = query.order_by(HardwareSupport.support_end_date.asc()).all()
    changed = False
    for item in items:
        old = item.status
        update_status(item)
        changed = changed or old != item.status
    if changed:
        db.commit()
    return items


def software_query(db: Session, q: str | None = None, status: str | None = None, responsible: str | None = None):
    query = db.query(SoftwareLicense)
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(
            SoftwareLicense.location.ilike(pattern), SoftwareLicense.code.ilike(pattern), SoftwareLicense.certificate_number.ilike(pattern),
            SoftwareLicense.contract_number.ilike(pattern), SoftwareLicense.contractor.ilike(pattern),
            SoftwareLicense.responsible.ilike(pattern), SoftwareLicense.comment.ilike(pattern),
        ))
    if status:
        query = query.filter(SoftwareLicense.status == status)
    if responsible:
        query = query.filter(SoftwareLicense.responsible == responsible)
    items = query.order_by(SoftwareLicense.cert_end_date.asc()).all()
    changed = False
    for item in items:
        old = item.status
        update_status(item)
        changed = changed or old != item.status
    if changed:
        db.commit()
    return items


def stats_for(items):
    today = date.today()
    return {
        "total": len(items),
        "expired": sum(1 for x in items if x.days_left(today) < 0),
        "urgent": sum(1 for x in items if 0 <= x.days_left(today) <= 7),
        "critical": sum(1 for x in items if 8 <= x.days_left(today) <= 30),
        "warning": sum(1 for x in items if 31 <= x.days_left(today) <= 60),
    }


def distinct_responsibles(db: Session, model):
    """Return unique responsible names for models with different field names.

HardwareSupport, SoftwareLicense and BackupObject use `responsible`;
DataProtectionPolicy uses `responsible_person`.
"""
    field = getattr(model, "responsible", None)
    if field is None:
        field = getattr(model, "responsible_person", None)
    if field is None:
        return []
    values = db.query(field).distinct().order_by(field).all()
    return [x[0] for x in values if x[0]]


DP_TYPE_LABELS = {
    "filesystem": "Filesystem",
    "internal_database": "Internal Database",
    "ms_sql_server": "MS SQL Server",
    "virtual_environment": "Virtual Environment",
}


def normalize_backup_type(value: str | None) -> str:
    value = str(value or "filesystem").strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "mssql": "ms_sql_server",
        "ms_sql": "ms_sql_server",
        "sql_server": "ms_sql_server",
        "virtual": "virtual_environment",
        "vm": "virtual_environment",
        "internal_db": "internal_database",
        "database": "internal_database",
    }
    return aliases.get(value, value if value in DP_TYPE_LABELS else "filesystem")


def parse_volume_gb(value) -> int | None:
    if value in (None, ""):
        return None
    raw = str(value).strip().lower().replace(",", ".")
    multiplier = 1
    if "tb" in raw or "тб" in raw:
        multiplier = 1024
    raw = re.sub(r"[^0-9.]", "", raw)
    if not raw:
        return None
    try:
        return int(round(float(raw) * multiplier))
    except Exception:
        return None


def format_volume(value: int | None) -> str:
    if value is None:
        return ""
    if value >= 1024:
        tb = value / 1024
        return f"{tb:.2f} ТБ".replace(".00", "")
    return f"{value} ГБ"


def dp_query(db: Session, q: str | None = None, backup_type: str | None = None, responsible: str | None = None):
    query = db.query(DataProtectionPolicy)
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(
            DataProtectionPolicy.system_name.ilike(pattern),
            DataProtectionPolicy.reserved_information.ilike(pattern),
            DataProtectionPolicy.procedure_frequency.ilike(pattern),
            DataProtectionPolicy.retention_period.ilike(pattern),
            DataProtectionPolicy.notes.ilike(pattern),
            DataProtectionPolicy.responsible_person.ilike(pattern),
        ))
    if backup_type:
        query = query.filter(DataProtectionPolicy.backup_type == normalize_backup_type(backup_type))
    if responsible:
        query = query.filter(DataProtectionPolicy.responsible_person == responsible)
    return query.order_by(DataProtectionPolicy.backup_type.asc(), DataProtectionPolicy.row_number.asc().nullslast(), DataProtectionPolicy.system_name.asc()).all()


def dp_grouped(items: list[DataProtectionPolicy]) -> list[dict]:
    result = {key: {"type": key, "label": label, "items": [], "count": 0, "volume_gb": 0, "volume_label": "0 ГБ"} for key, label in DP_TYPE_LABELS.items()}
    for item in items:
        key = normalize_backup_type(item.backup_type)
        row = result.setdefault(key, {"type": key, "label": DP_TYPE_LABELS.get(key, key), "items": [], "count": 0, "volume_gb": 0, "volume_label": "0 ГБ"})
        row["items"].append(item)
        row["count"] += 1
        row["volume_gb"] += item.max_volume_gb or 0
    for row in result.values():
        row["volume_label"] = format_volume(row["volume_gb"])
    return [row for row in result.values() if row["count"] > 0]


def dp_stats(items: list[DataProtectionPolicy]) -> dict:
    total = len(items)
    volume = sum(x.max_volume_gb or 0 for x in items)
    full_count = sum(1 for x in items if x.procedure_frequency and "full" in x.procedure_frequency.lower())
    incr_count = sum(1 for x in items if x.procedure_frequency and ("incr" in x.procedure_frequency.lower() or "increment" in x.procedure_frequency.lower()))
    no_responsible = sum(1 for x in items if not x.responsible_person)
    return {"total": total, "volume_gb": volume, "volume_label": format_volume(volume), "types": len({x.backup_type for x in items}) if items else 0, "full": full_count, "incremental": incr_count, "no_responsible": no_responsible}


def dp_type_stats(items: list[DataProtectionPolicy]) -> list[dict]:
    rows = []
    for group in dp_grouped(items):
        rows.append({"type": group["type"], "label": group["label"], "count": group["count"], "volume_gb": group["volume_gb"], "volume_label": group["volume_label"]})
    rows.sort(key=lambda x: x["volume_gb"], reverse=True)
    return rows


def dp_to_api(item: DataProtectionPolicy) -> dict:
    return {
        "id": item.id,
        "row_number": item.row_number,
        "backup_type": item.backup_type,
        "backup_type_label": DP_TYPE_LABELS.get(item.backup_type, item.backup_type),
        "system_name": item.system_name,
        "reserved_information": item.reserved_information,
        "max_volume_gb": item.max_volume_gb,
        "procedure_frequency": item.procedure_frequency,
        "retention_period": item.retention_period,
        "notes": item.notes,
        "responsible_person": item.responsible_person,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def backup_query(db: Session, q: str | None = None, status: str | None = None, responsible: str | None = None):
    query = db.query(BackupObject)
    if q:
        pattern = f"%{q}%"
        query = query.filter(or_(
            BackupObject.object_name.ilike(pattern), BackupObject.object_type.ilike(pattern),
            BackupObject.platform.ilike(pattern), BackupObject.storage.ilike(pattern),
            BackupObject.policy_name.ilike(pattern), BackupObject.responsible.ilike(pattern),
            BackupObject.comment.ilike(pattern),
        ))
    if status:
        query = query.filter(BackupObject.status == status)
    if responsible:
        query = query.filter(BackupObject.responsible == responsible)
    return query.order_by(BackupObject.last_backup_at.desc().nullslast(), BackupObject.object_name.asc()).all()


def backup_stats(items: list[BackupObject]) -> dict:
    total = len(items)
    success = sum(1 for x in items if x.status == "success")
    warning = sum(1 for x in items if x.status == "warning")
    failed = sum(1 for x in items if x.status == "failed")
    running = sum(1 for x in items if x.status == "running")
    unknown = sum(1 for x in items if x.status == "unknown")
    health = round((success / total) * 100, 1) if total else 0
    return {"total": total, "success": success, "warning": warning, "failed": failed, "running": running, "unknown": unknown, "health": health}


def backup_platform_stats(items: list[BackupObject]) -> list[dict]:
    result = {}
    for item in items:
        row = result.setdefault(item.platform or "manual", {"platform": item.platform or "manual", "total": 0, "success": 0, "warning": 0, "failed": 0, "running": 0, "unknown": 0})
        row["total"] += 1
        if item.status in row:
            row[item.status] += 1
    rows = list(result.values())
    for row in rows:
        row["success_percent"] = round(row["success"] / row["total"] * 100, 1) if row["total"] else 0
    rows.sort(key=lambda x: x["total"], reverse=True)
    return rows


def backup_type_stats(items: list[BackupObject]) -> list[dict]:
    result = {}
    for item in items:
        key = item.object_type or "other"
        result[key] = result.get(key, 0) + 1
    return [{"type": k, "count": v} for k, v in sorted(result.items(), key=lambda x: x[1], reverse=True)]


def parse_datetime_local(value) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text_value = str(value).strip()
    if not text_value:
        return None
    # HTML datetime-local приходит как 2026-05-22T03:00
    return datetime.fromisoformat(text_value[:16])


def format_dt_for_input(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%dT%H:%M") if value else ""


def backup_to_api(item: BackupObject) -> dict:
    return {
        "id": item.id,
        "object_name": item.object_name,
        "object_type": item.object_type,
        "platform": item.platform,
        "last_backup_at": item.last_backup_at,
        "next_backup_at": item.next_backup_at,
        "size_gb": item.size_gb,
        "storage": item.storage,
        "retention_days": item.retention_days,
        "duration_min": item.duration_min,
        "policy_name": item.policy_name,
        "responsible": item.responsible,
        "status": item.status,
        "comment": item.comment,
        "hours_since_backup": item.hours_since_backup(),
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }



@app.get("/admin/backups", response_class=HTMLResponse)
def backups_page(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    return templates.TemplateResponse(
        request=request,
        name="backups.html",
        context={
            "app_name": settings.app_name,
            "backups": list_db_backups(),
            "backup_dir": str(get_backup_dir()),
            "database_path": str(get_sqlite_db_path()),
            "current_user": current_user,
        },
    )


@app.post("/admin/backups/create")
def create_backup(db: Session = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    backup_path = create_sqlite_backup(current_user.username)
    log_change(db, "system", None, "backup", f"Создана резервная копия базы: {backup_path.name}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/admin/backups", status_code=303)


@app.get("/admin/backups/{filename}/download")
def download_backup(filename: str, current_user: User = Depends(require_role("admin"))):
    path = validate_backup_filename(filename)
    return FileResponse(path=str(path), media_type="application/octet-stream", filename=path.name)


@app.post("/admin/backups/{filename}/delete")
def delete_backup(filename: str, db: Session = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    path = validate_backup_filename(filename)
    path.unlink()
    log_change(db, "system", None, "backup_delete", f"Удалена резервная копия базы: {filename}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/admin/backups", status_code=303)



@app.get("/backup-monitor", response_class=HTMLResponse)
def backup_monitor_page(request: Request, q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    if settings.auth_enabled and current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    items = backup_query(db, q, status, responsible)
    all_items = backup_query(db)
    return templates.TemplateResponse(
        request=request,
        name="backup_monitor.html",
        context={
            "app_name": settings.app_name,
            "items": items,
            "all_items": all_items,
            "stats": backup_stats(all_items),
            "platform_stats": backup_platform_stats(all_items),
            "type_stats": backup_type_stats(all_items),
            "selected": {"q": q or "", "status": status or "", "responsible": responsible or ""},
            "responsibles": distinct_responsibles(db, BackupObject),
            "statuses": ["success", "warning", "failed", "running", "unknown"],
            "current_user": current_user,
            "can_edit": user_can(current_user, "editor"),
            "can_manage": user_can(current_user, "manager"),
            "can_admin": user_can(current_user, "admin"),
            "format_dt_for_input": format_dt_for_input,
        },
    )


@app.post("/backup-monitor")
def create_backup_object(
    object_name: str = Form("-"), object_type: str = Form("vm"), platform: str = Form("manual"),
    last_backup_at: str | None = Form(None), next_backup_at: str | None = Form(None), size_gb: str | None = Form(None),
    storage: str | None = Form(None), retention_days: str | None = Form(None), duration_min: str | None = Form(None),
    policy_name: str | None = Form(None), responsible: str | None = Form(None), status: str = Form("success"),
    comment: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = BackupObject(
        object_name=object_name or "-", object_type=object_type or "vm", platform=platform or "manual",
        last_backup_at=parse_datetime_local(last_backup_at), next_backup_at=parse_datetime_local(next_backup_at),
        size_gb=parse_int(size_gb), storage=storage or None, retention_days=parse_int(retention_days),
        duration_min=parse_int(duration_min), policy_name=policy_name or None, responsible=responsible or None,
        status=status or "success", comment=comment or None,
    )
    db.add(item)
    db.flush()
    log_change(db, "backup_monitor", item.id, "create", f"Добавлен объект резервного копирования: {item.object_name}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/backup-monitor", status_code=303)


@app.post("/backup-monitor/{item_id}/update")
def update_backup_object(
    item_id: int,
    object_name: str = Form("-"), object_type: str = Form("vm"), platform: str = Form("manual"),
    last_backup_at: str | None = Form(None), next_backup_at: str | None = Form(None), size_gb: str | None = Form(None),
    storage: str | None = Form(None), retention_days: str | None = Form(None), duration_min: str | None = Form(None),
    policy_name: str | None = Form(None), responsible: str | None = Form(None), status: str = Form("success"),
    comment: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = db.get(BackupObject, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Backup object not found")
    item.object_name = object_name or "-"
    item.object_type = object_type or "vm"
    item.platform = platform or "manual"
    item.last_backup_at = parse_datetime_local(last_backup_at)
    item.next_backup_at = parse_datetime_local(next_backup_at)
    item.size_gb = parse_int(size_gb)
    item.storage = storage or None
    item.retention_days = parse_int(retention_days)
    item.duration_min = parse_int(duration_min)
    item.policy_name = policy_name or None
    item.responsible = responsible or None
    item.status = status or "success"
    item.comment = comment or None
    log_change(db, "backup_monitor", item.id, "update", f"Обновлён объект резервного копирования: {item.object_name}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/backup-monitor", status_code=303)


@app.post("/backup-monitor/{item_id}/delete")
def delete_backup_object(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    item = db.get(BackupObject, item_id)
    if item:
        log_change(db, "backup_monitor", item.id, "delete", f"Удалён объект резервного копирования: {item.object_name}", actor=current_user.username)
        db.delete(item)
        db.commit()
    return RedirectResponse(url="/backup-monitor", status_code=303)


@app.get("/api/backup-monitor")
def api_backup_monitor(request: Request, q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    return [backup_to_api(x) for x in backup_query(db, q, status, responsible)]


@app.get("/api/backup-monitor/summary")
def api_backup_monitor_summary(request: Request, db: Session = Depends(get_db)):
    require_api_key(request)
    items = backup_query(db)
    return {"stats": backup_stats(items), "platforms": backup_platform_stats(items), "types": backup_type_stats(items)}


@app.get("/data-protection", response_class=HTMLResponse)
def data_protection_page(request: Request, q: str | None = None, backup_type: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    if settings.auth_enabled and current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    items = dp_query(db, q, backup_type, responsible)
    all_items = dp_query(db)
    return templates.TemplateResponse(
        request=request,
        name="data_protection.html",
        context={
            "app_name": settings.app_name,
            "items": items,
            "all_items": all_items,
            "groups": dp_grouped(items),
            "stats": dp_stats(all_items),
            "type_stats": dp_type_stats(all_items),
            "backup_types": DP_TYPE_LABELS,
            "selected": {"q": q or "", "backup_type": backup_type or "", "responsible": responsible or ""},
            "responsibles": distinct_responsibles(db, DataProtectionPolicy),
            "current_user": current_user,
            "can_edit": user_can(current_user, "editor"),
            "can_manage": user_can(current_user, "manager"),
            "can_admin": user_can(current_user, "admin"),
            "format_volume": format_volume,
        },
    )


@app.post("/data-protection")
def create_data_protection_policy(
    row_number: str | None = Form(None), backup_type: str = Form("filesystem"), system_name: str = Form("-"),
    reserved_information: str | None = Form(None), max_volume_gb: str | None = Form(None), procedure_frequency: str | None = Form(None),
    retention_period: str | None = Form(None), notes: str | None = Form(None), responsible_person: str | None = Form(None),
    db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = DataProtectionPolicy(
        row_number=parse_int(row_number), backup_type=normalize_backup_type(backup_type), system_name=system_name or "-",
        reserved_information=reserved_information or None, max_volume_gb=parse_volume_gb(max_volume_gb),
        procedure_frequency=procedure_frequency or None, retention_period=retention_period or None, notes=notes or None,
        responsible_person=responsible_person or None,
    )
    db.add(item)
    db.flush()
    log_change(db, "data_protection", item.id, "create", f"Добавлена политика резервирования: {item.system_name}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


@app.post("/data-protection/{item_id}/update")
def update_data_protection_policy(
    item_id: int, row_number: str | None = Form(None), backup_type: str = Form("filesystem"), system_name: str = Form("-"),
    reserved_information: str | None = Form(None), max_volume_gb: str | None = Form(None), procedure_frequency: str | None = Form(None),
    retention_period: str | None = Form(None), notes: str | None = Form(None), responsible_person: str | None = Form(None),
    db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = db.get(DataProtectionPolicy, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Data protection policy not found")
    item.row_number = parse_int(row_number)
    item.backup_type = normalize_backup_type(backup_type)
    item.system_name = system_name or "-"
    item.reserved_information = reserved_information or None
    item.max_volume_gb = parse_volume_gb(max_volume_gb)
    item.procedure_frequency = procedure_frequency or None
    item.retention_period = retention_period or None
    item.notes = notes or None
    item.responsible_person = responsible_person or None
    log_change(db, "data_protection", item.id, "update", f"Обновлена политика резервирования: {item.system_name}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


@app.post("/data-protection/{item_id}/delete")
def delete_data_protection_policy(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    item = db.get(DataProtectionPolicy, item_id)
    if item:
        log_change(db, "data_protection", item.id, "delete", f"Удалена политика резервирования: {item.system_name}", actor=current_user.username)
        db.delete(item)
        db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


@app.get("/api/data-protection/plan")
def api_data_protection_plan(request: Request, q: str | None = None, backup_type: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    return [dp_to_api(x) for x in dp_query(db, q, backup_type, responsible)]


@app.get("/api/data-protection/summary")
def api_data_protection_summary(request: Request, db: Session = Depends(get_db)):
    require_api_key(request)
    items = dp_query(db)
    return {"stats": dp_stats(items), "types": dp_type_stats(items)}


@app.post("/data-protection/clear")
def clear_data_protection_policies(db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    deleted = db.query(DataProtectionPolicy).delete(synchronize_session=False)
    log_change(db, "data_protection", None, "clear", f"Очищена таблица политик резервирования. Удалено записей: {deleted}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


def clean_cell(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def detect_docx_backup_type(row_values: list[str], current_type: str) -> str:
    joined = " ".join(v.lower() for v in row_values if v)
    if "filesystem" in joined or "file system" in joined or "файлов" in joined:
        return "filesystem"
    if "internal database" in joined or "internal db" in joined or "встроенн" in joined:
        return "internal_database"
    if "ms sql" in joined or "mssql" in joined or "sql server" in joined:
        return "ms_sql_server"
    if "virtual environment" in joined or "virtual" in joined or "vmware" in joined or "виртуаль" in joined:
        return "virtual_environment"
    return current_type


def is_docx_header_row(row_values: list[str]) -> bool:
    joined = " ".join(v.lower() for v in row_values if v)
    header_words = ["наименование", "информац", "периодич", "срок", "ответствен", "резерв"]
    return sum(1 for word in header_words if word in joined) >= 3


def row_looks_like_group(row_values: list[str]) -> bool:
    non_empty = [v for v in row_values if v]
    if len(non_empty) == 1:
        value = non_empty[0].lower()
        return any(marker in value for marker in ["filesystem", "internal", "sql", "virtual", "файлов", "виртуаль"])
    return False


@app.get("/data-protection/export.xlsx")
def export_data_protection_xlsx(db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "data_protection_plan"
    headers = ["№", "Тип backup", "Наименование информационной системы", "Информация, подлежащая резервированию", "Максимальный объём, ГБ", "Периодичность проведения процедуры", "Срок хранения информации", "Примечания", "Ответственный работник"]
    ws.append(headers)
    for item in dp_query(db):
        ws.append([item.row_number, DP_TYPE_LABELS.get(item.backup_type, item.backup_type), item.system_name, item.reserved_information, item.max_volume_gb, item.procedure_frequency, item.retention_period, item.notes, item.responsible_person])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 55)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="data_protection_plan_{date.today().isoformat()}.xlsx"'})


@app.post("/data-protection/import.xlsx")
async def import_data_protection_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload .xlsx file")
    wb = load_workbook(BytesIO(await file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if not any(row):
            continue
        item = DataProtectionPolicy(
            row_number=parse_int(row[0] if len(row) > 0 else None),
            backup_type=normalize_backup_type(row[1] if len(row) > 1 else None),
            system_name=str(row[2] if len(row) > 2 and row[2] is not None else "-").strip(),
            reserved_information=str(row[3]).strip() if len(row) > 3 and row[3] is not None else None,
            max_volume_gb=parse_volume_gb(row[4] if len(row) > 4 else None),
            procedure_frequency=str(row[5]).strip() if len(row) > 5 and row[5] is not None else None,
            retention_period=str(row[6]).strip() if len(row) > 6 and row[6] is not None else None,
            notes=str(row[7]).strip() if len(row) > 7 and row[7] is not None else None,
            responsible_person=str(row[8]).strip() if len(row) > 8 and row[8] is not None else None,
        )
        db.add(item)
    log_change(db, "data_protection", None, "import", f"Импортирован план резервирования из Excel: {file.filename}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


@app.post("/data-protection/import.docx")
async def import_data_protection_docx(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    if Document is None:
        raise HTTPException(status_code=500, detail="python-docx is not installed. Run: pip3 install python-docx")
    if not file.filename.lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="Upload .docx file")

    document = Document(BytesIO(await file.read()))
    current_type = "filesystem"
    imported = 0

    for table in document.tables:
        for row in table.rows:
            values = [clean_cell(cell.text) for cell in row.cells]
            if not any(values):
                continue

            detected_type = detect_docx_backup_type(values, current_type)
            if detected_type != current_type or row_looks_like_group(values):
                current_type = detected_type
                if row_looks_like_group(values):
                    continue

            if is_docx_header_row(values):
                continue

            # Ожидаемый порядок из Word:
            # № | наименование ИС | информация | max GB | периодичность | срок хранения | примечания | ответственный
            row_number = values[0] if len(values) > 0 else None
            system_name = values[1] if len(values) > 1 else ""
            reserved_information = values[2] if len(values) > 2 else ""
            max_volume = values[3] if len(values) > 3 else ""
            procedure_frequency = values[4] if len(values) > 4 else ""
            retention_period = values[5] if len(values) > 5 else ""
            notes = values[6] if len(values) > 6 else ""
            responsible_person = values[7] if len(values) > 7 else ""

            if not system_name and not reserved_information:
                continue

            item = DataProtectionPolicy(
                row_number=parse_int(row_number),
                backup_type=current_type,
                system_name=system_name or "-",
                reserved_information=reserved_information or None,
                max_volume_gb=parse_volume_gb(max_volume),
                procedure_frequency=procedure_frequency or None,
                retention_period=retention_period or None,
                notes=notes or None,
                responsible_person=responsible_person or None,
            )
            db.add(item)
            imported += 1

    log_change(db, "data_protection", None, "import", f"Импортирован план резервирования из Word: {file.filename}. Записей: {imported}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/data-protection", status_code=303)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if not settings.auth_enabled:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse(request=request, name="login.html", context={"app_name": settings.app_name, "error": None})


@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    if user is None or not user.is_active or not verify_password(password, user.password_hash):
        return templates.TemplateResponse(request=request, name="login.html", context={"app_name": settings.app_name, "error": "Неверный логин или пароль"}, status_code=401)
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(SESSION_COOKIE, create_session_token(user.id), httponly=True, samesite="lax", max_age=60 * 60 * 12)
    return response


@app.post("/logout")
def logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/admin/users", response_class=HTMLResponse)
def users_page(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_role("admin"))):
    users = db.query(User).order_by(User.username).all()
    return templates.TemplateResponse(request=request, name="users.html", context={"app_name": settings.app_name, "users": users, "roles": ROLE_LABELS, "current_user": current_user})


@app.post("/admin/users")
def create_user(
    username: str = Form(...), password: str = Form(...), full_name: str | None = Form(None),
    email: str | None = Form(None), role: str = Form("viewer"), is_active: bool = Form(False),
    db: Session = Depends(get_db), current_user: User = Depends(require_role("admin")),
):
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=400, detail="Invalid role")
    existing = db.query(User).filter(User.username == username).first()
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    db.add(User(username=username, password_hash=hash_password(password), full_name=full_name or None, email=email or None, role=role, is_active=is_active))
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


@app.post("/admin/users/{user_id}/update")
def update_user(
    user_id: int, full_name: str | None = Form(None), email: str | None = Form(None), role: str = Form("viewer"),
    is_active: bool = Form(False), password: str | None = Form(None), db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    if role not in ROLE_LABELS:
        raise HTTPException(status_code=400, detail="Invalid role")
    user = db.get(User, user_id)
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    user.full_name = full_name or None
    user.email = email or None
    user.role = role
    user.is_active = is_active
    if password:
        user.password_hash = hash_password(password)
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


@app.get("/", response_class=HTMLResponse)
def index(request: Request, section: str = "hardware", q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    if settings.auth_enabled and current_user is None:
        return RedirectResponse(url="/login", status_code=303)
    if section not in {"hardware", "software"}:
        section = "hardware"
    hw_all = hardware_query(db)
    sw_all = software_query(db)
    items = hardware_query(db, q, status, responsible) if section == "hardware" else software_query(db, q, status, responsible)
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "app_name": settings.app_name,
            "section": section,
            "items": items,
            "today": date.today(),
            "stats": stats_for(hw_all if section == "hardware" else sw_all),
            "global_stats": {"hardware": stats_for(hw_all), "software": stats_for(sw_all)},
            "selected": {"q": q or "", "status": status or "", "responsible": responsible or ""},
            "responsibles": distinct_responsibles(db, HardwareSupport if section == "hardware" else SoftwareLicense),
            "statuses": ["active", "warning", "critical", "urgent", "expired"],
            "current_user": current_user,
            "can_edit": user_can(current_user, "editor"),
            "can_manage": user_can(current_user, "manager"),
            "can_admin": user_can(current_user, "admin"),
        },
    )


@app.post("/hardware")
def create_hardware(
    dzo: str = Form("-"), serial_number: str = Form("-"), equipment_model: str = Form("-"), delivery_year: str | None = Form(None),
    support_end_date: str = Form(...), purchase_period: str | None = Form(None), contract_number: str | None = Form(None),
    contractor: str | None = Form(None), responsible: str | None = Form(None), auto_renewal: bool = Form(False),
    comment: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    end = parse_date(support_end_date)
    if not end:
        raise HTTPException(status_code=400, detail="Дата окончания ТП обязательна")
    item = HardwareSupport(
        dzo=dzo or "-", serial_number=serial_number or "-", equipment_model=equipment_model or "-", delivery_year=parse_int(delivery_year),
        support_end_date=end, purchase_period=purchase_period or None, contract_number=contract_number or None,
        contractor=contractor or None, responsible=responsible or None, auto_renewal=auto_renewal, comment=comment or None,
    )
    update_status(item)
    db.add(item)
    db.flush()
    log_change(db, "hardware", item.id, "create", f"Добавлена ТП оборудования: {item.equipment_model} / {item.serial_number}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/?section=hardware", status_code=303)


@app.post("/software")
def create_software(
    location: str | None = Form(None), code: str = Form("-"), quantity: int = Form(1), cert_start_date: str | None = Form(None), cert_end_date: str = Form(...),
    certificate_number: str | None = Form(None), comment: str | None = Form(None), contract_number: str | None = Form(None),
    contractor: str | None = Form(None), responsible: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    end = parse_date(cert_end_date)
    if not end:
        raise HTTPException(status_code=400, detail="Дата окончания сертификата обязательна")
    item = SoftwareLicense(
        location=location or None, code=code or "-", quantity=quantity or 1, cert_start_date=parse_date(cert_start_date), cert_end_date=end,
        certificate_number=certificate_number or None, comment=comment or None, contract_number=contract_number or None,
        contractor=contractor or None, responsible=responsible or None,
    )
    update_status(item)
    db.add(item)
    db.flush()
    log_change(db, "software", item.id, "create", f"Добавлена лицензия ПО: {item.code}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/?section=software", status_code=303)


@app.post("/hardware/{item_id}/delete")
def delete_hardware(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    item = db.get(HardwareSupport, item_id)
    if item:
        log_change(db, "hardware", item.id, "delete", f"Удалена ТП оборудования: {item.equipment_model} / {item.serial_number}", actor=current_user.username)
        db.delete(item)
        db.commit()
    return RedirectResponse(url="/?section=hardware", status_code=303)


@app.post("/software/{item_id}/delete")
def delete_software(item_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    item = db.get(SoftwareLicense, item_id)
    if item:
        log_change(db, "software", item.id, "delete", f"Удалена лицензия ПО: {item.code}", actor=current_user.username)
        db.delete(item)
        db.commit()
    return RedirectResponse(url="/?section=software", status_code=303)


@app.post("/hardware/{item_id}/update")
def update_hardware(
    item_id: int,
    dzo: str = Form("-"), serial_number: str = Form("-"), equipment_model: str = Form("-"), delivery_year: str | None = Form(None),
    support_end_date: str = Form(...), purchase_period: str | None = Form(None), contract_number: str | None = Form(None),
    contractor: str | None = Form(None), responsible: str | None = Form(None), auto_renewal: bool = Form(False),
    comment: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = db.get(HardwareSupport, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Hardware support record not found")
    end = parse_date(support_end_date)
    if not end:
        raise HTTPException(status_code=400, detail="Дата окончания ТП обязательна")
    item.dzo = dzo or "-"
    item.serial_number = serial_number or "-"
    item.equipment_model = equipment_model or "-"
    item.delivery_year = parse_int(delivery_year)
    item.support_end_date = end
    item.purchase_period = purchase_period or None
    item.contract_number = contract_number or None
    item.contractor = contractor or None
    item.responsible = responsible or None
    item.auto_renewal = auto_renewal
    item.comment = comment or None
    update_status(item)
    log_change(db, "hardware", item.id, "update", f"Обновлена ТП оборудования: {item.equipment_model} / {item.serial_number}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/?section=hardware", status_code=303)


@app.post("/software/{item_id}/update")
def update_software(
    item_id: int,
    location: str | None = Form(None), code: str = Form("-"), quantity: int = Form(1), cert_start_date: str | None = Form(None), cert_end_date: str = Form(...),
    certificate_number: str | None = Form(None), comment: str | None = Form(None), contract_number: str | None = Form(None),
    contractor: str | None = Form(None), responsible: str | None = Form(None), db: Session = Depends(get_db), current_user: User = Depends(require_role("editor")),
):
    item = db.get(SoftwareLicense, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Software license record not found")
    end = parse_date(cert_end_date)
    if not end:
        raise HTTPException(status_code=400, detail="Дата окончания сертификата обязательна")
    item.location = location or None
    item.code = code or "-"
    item.quantity = quantity or 1
    item.cert_start_date = parse_date(cert_start_date)
    item.cert_end_date = end
    item.certificate_number = certificate_number or None
    item.comment = comment or None
    item.contract_number = contract_number or None
    item.contractor = contractor or None
    item.responsible = responsible or None
    update_status(item)
    log_change(db, "software", item.id, "update", f"Обновлена лицензия ПО: {item.code}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url="/?section=software", status_code=303)


@app.get("/history", response_class=HTMLResponse)
def history_page(request: Request, db: Session = Depends(get_db), current_user: User = Depends(require_login)):
    history = db.query(ChangeLog).order_by(ChangeLog.created_at.desc()).limit(200).all()
    return templates.TemplateResponse(request=request, name="history.html", context={"history": history, "app_name": settings.app_name, "current_user": current_user, "can_admin": user_can(current_user, "admin") })


def hardware_to_api(item: HardwareSupport) -> dict:
    today = date.today()
    return {
        "id": item.id,
        "section": "hardware",
        "section_name": "Техподдержка оборудования",
        "number": None,
        "dzo": item.dzo,
        "serial_number": item.serial_number,
        "equipment_model": item.equipment_model,
        "delivery_year": item.delivery_year,
        "support_end_date": item.support_end_date,
        "purchase_period": item.purchase_period,
        "contract_number": item.contract_number,
        "contractor": item.contractor,
        "responsible": item.responsible,
        "auto_renewal": item.auto_renewal,
        "comment": item.comment,
        "days_left": item.days_left(today),
        "status": item.status,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def software_to_api(item: SoftwareLicense) -> dict:
    today = date.today()
    return {
        "id": item.id,
        "section": "software",
        "section_name": "Лицензии ПО",
        "number": None,
        "location": item.location,
        "code": item.code,
        "quantity": item.quantity,
        "cert_start_date": item.cert_start_date,
        "cert_end_date": item.cert_end_date,
        "certificate_number": item.certificate_number,
        "comment": item.comment,
        "contract_number": item.contract_number,
        "contractor": item.contractor,
        "responsible": item.responsible,
        "days_left": item.days_left(today),
        "status": item.status,
        "created_at": item.created_at,
        "updated_at": item.updated_at,
    }


def combined_api_items(db: Session, days: int | None = None) -> list[dict]:
    result = []
    for item in hardware_query(db):
        if days is None or item.days_left(date.today()) <= days:
            result.append(hardware_to_api(item))
    for item in software_query(db):
        if days is None or item.days_left(date.today()) <= days:
            result.append(software_to_api(item))
    result.sort(key=lambda x: x.get("days_left", 999999))
    for idx, item in enumerate(result, 1):
        item["number"] = idx
    return result



@app.get("/api/licenses")
def api_all_licenses(request: Request, section: str | None = None, q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    """Совместимый API: возвращает записи обоих разделов или выбранного section=hardware/software."""
    if section == "hardware":
        return [hardware_to_api(x) for x in hardware_query(db, q, status, responsible)]
    if section == "software":
        return [software_to_api(x) for x in software_query(db, q, status, responsible)]
    return combined_api_items(db)


@app.get("/api/licenses/expiring")
def api_expiring_licenses(request: Request, days: int = 30, section: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    today = date.today()
    if section == "hardware":
        return [hardware_to_api(x) for x in hardware_query(db) if x.days_left(today) <= days]
    if section == "software":
        return [software_to_api(x) for x in software_query(db) if x.days_left(today) <= days]
    return combined_api_items(db, days=days)


@app.get("/api/hardware-support")
def api_hardware_support(request: Request, q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    return [hardware_to_api(x) for x in hardware_query(db, q, status, responsible)]


@app.get("/api/hardware-support/{item_id}")
def api_hardware_support_item(request: Request, item_id: int, db: Session = Depends(get_db)):
    require_api_key(request)
    item = db.get(HardwareSupport, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Hardware support record not found")
    update_status(item)
    db.commit()
    return hardware_to_api(item)


@app.get("/api/software-licenses")
def api_software_licenses(request: Request, q: str | None = None, status: str | None = None, responsible: str | None = None, db: Session = Depends(get_db)):
    require_api_key(request)
    return [software_to_api(x) for x in software_query(db, q, status, responsible)]


@app.get("/api/software-licenses/{item_id}")
def api_software_license_item(request: Request, item_id: int, db: Session = Depends(get_db)):
    require_api_key(request)
    item = db.get(SoftwareLicense, item_id)
    if item is None:
        raise HTTPException(status_code=404, detail="Software license record not found")
    update_status(item)
    db.commit()
    return software_to_api(item)

@app.get("/api/monitoring/summary")
def api_monitoring_summary(request: Request, db: Session = Depends(get_db)):
    require_api_key(request)
    hw = hardware_query(db)
    sw = software_query(db)
    return {"hardware_support": stats_for(hw), "software_licenses": stats_for(sw), "backup_monitor": backup_stats(backup_query(db)), "data_protection": dp_stats(dp_query(db))}




@app.get("/api/zabbix/summary")
def api_zabbix_summary(request: Request, db: Session = Depends(get_db)):
    require_api_key(request)
    """Плоский JSON для Zabbix HTTP agent и JSONPath."""
    hw = stats_for(hardware_query(db))
    sw = stats_for(software_query(db))
    return {
        "hardware_total": hw["total"],
        "hardware_warning": hw["warning"],
        "hardware_critical": hw["critical"],
        "hardware_urgent": hw["urgent"],
        "hardware_expired": hw["expired"],
        "software_total": sw["total"],
        "software_warning": sw["warning"],
        "software_critical": sw["critical"],
        "software_urgent": sw["urgent"],
        "software_expired": sw["expired"],
        "total_expired": hw["expired"] + sw["expired"],
        "total_urgent": hw["urgent"] + sw["urgent"],
        "total_critical": hw["critical"] + sw["critical"],
        "total_warning": hw["warning"] + sw["warning"],
        "backup_total": backup_stats(backup_query(db))["total"],
        "backup_success": backup_stats(backup_query(db))["success"],
        "backup_warning": backup_stats(backup_query(db))["warning"],
        "backup_failed": backup_stats(backup_query(db))["failed"],
        "backup_running": backup_stats(backup_query(db))["running"],
        "backup_health": backup_stats(backup_query(db))["health"],
        "data_protection_total": dp_stats(dp_query(db))["total"],
        "data_protection_volume_gb": dp_stats(dp_query(db))["volume_gb"],
        "data_protection_no_responsible": dp_stats(dp_query(db))["no_responsible"],
    }

@app.get("/metrics", response_class=PlainTextResponse)
def metrics(request: Request, db: Session = Depends(get_db)):
    require_api_key(request)
    hw = stats_for(hardware_query(db))
    sw = stats_for(software_query(db))
    lines = []
    for section, data in {"hardware": hw, "software": sw}.items():
        for key, value in data.items():
            lines.append(f'license_monitor_{key}{{section="{section}"}} {value}')
    backup = backup_stats(backup_query(db))
    for key, value in backup.items():
        lines.append(f'license_monitor_backup_{key} {value}')
    dp = dp_stats(dp_query(db))
    for key, value in dp.items():
        if isinstance(value, (int, float)):
            lines.append(f'license_monitor_data_protection_{key} {value}')
    return "\n".join(lines) + "\n"


@app.get("/export.xlsx")
def export_xlsx(section: str = "hardware", db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    wb = Workbook()
    ws = wb.active
    today = date.today()
    if section == "software":
        ws.title = "software_licenses"
        headers = ["№", "Локация", "Код", "Кол-во", "Дата начала сертификата", "Дата окончания сертификата", "Номер сертификата", "Комментарий", "Номер договора", "Контрагент", "Ответственный", "Осталось дней", "Статус"]
        ws.append(headers)
        for idx, item in enumerate(software_query(db), 1):
            ws.append([idx, item.location, item.code, item.quantity, item.cert_start_date, item.cert_end_date, item.certificate_number, item.comment, item.contract_number, item.contractor, item.responsible, item.days_left(today), item.status])
        filename = f"software_licenses_{today.isoformat()}.xlsx"
    else:
        ws.title = "hardware_support"
        headers = ["№", "ДЗО", "Серийный номер", "Модель оборудования", "Год поставки", "Текущее состояние ТП (дата окончания)", "Срок на который закупается ТП", "Договор", "Контрагент", "Ответственный", "Автопродление", "Комментарий", "Осталось дней", "Статус"]
        ws.append(headers)
        for idx, item in enumerate(hardware_query(db), 1):
            ws.append([idx, item.dzo, item.serial_number, item.equipment_model, item.delivery_year, item.support_end_date, item.purchase_period, item.contract_number, item.contractor, item.responsible, "да" if item.auto_renewal else "нет", item.comment, item.days_left(today), item.status])
        filename = f"hardware_support_{today.isoformat()}.xlsx"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 45)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/import.xlsx")
async def import_xlsx(section: str = Form("hardware"), file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(require_role("manager"))):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Upload .xlsx file")
    wb = load_workbook(BytesIO(await file.read()), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return RedirectResponse(url=f"/?section={section}", status_code=303)
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    index = {name: i for i, name in enumerate(headers)}

    def val(row, name, default=None):
        i = index.get(name)
        return row[i] if i is not None and i < len(row) else default

    imported = 0
    for row in rows[1:]:
        if not row or not any(row):
            continue
        try:
            if section == "software":
                end = parse_date(val(row, "Дата окончания сертификата"))
                if not end:
                    continue
                item = SoftwareLicense(
                    location=str(val(row, "Локация") or "").strip() or None,
                    code=str(val(row, "Код", "-") or "-").strip(), quantity=parse_int(val(row, "Кол-во"), 1) or 1,
                    cert_start_date=parse_date(val(row, "Дата начала сертификата")), cert_end_date=end,
                    certificate_number=str(val(row, "Номер сертификата") or "").strip() or None,
                    comment=str(val(row, "Комментарий") or "").strip() or None,
                    contract_number=str(val(row, "Номер договора") or "").strip() or None,
                    contractor=str(val(row, "Контрагент") or "").strip() or None,
                    responsible=str(val(row, "Ответственный") or "").strip() or None,
                )
                update_status(item)
                db.add(item)
                imported += 1
            else:
                end = parse_date(val(row, "Текущее состояние ТП (дата окончания)"))
                if not end:
                    continue
                item = HardwareSupport(
                    dzo=str(val(row, "ДЗО", "-") or "-").strip(), serial_number=str(val(row, "Серийный номер", "-") or "-").strip(),
                    equipment_model=str(val(row, "Модель оборудования", "-") or "-").strip(), delivery_year=parse_int(val(row, "Год поставки")),
                    support_end_date=end, purchase_period=str(val(row, "Срок на который закупается ТП") or "").strip() or None,
                    contract_number=str(val(row, "Договор") or "").strip() or None, contractor=str(val(row, "Контрагент") or "").strip() or None,
                    responsible=str(val(row, "Ответственный") or "").strip() or None, auto_renewal=parse_bool(val(row, "Автопродление")),
                    comment=str(val(row, "Комментарий") or "").strip() or None,
                )
                update_status(item)
                db.add(item)
                imported += 1
        except Exception:
            continue
    log_change(db, section, None, "import", f"Импортировано строк из Excel: {imported}", actor=current_user.username)
    db.commit()
    return RedirectResponse(url=f"/?section={section}", status_code=303)


@app.get("/health")
def health():
    return {"status": "ok", "app": settings.app_name}
